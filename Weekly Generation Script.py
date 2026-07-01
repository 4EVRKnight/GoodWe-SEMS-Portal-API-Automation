"""
GoodWe SEMS Weekly Generation Automation  v12  (RATE-LIMIT SAFE + OFFLINE SHEET)
======================================================
Usage:
    python goodwe_automation.py              → normal run
    python goodwe_automation.py --preview    → show what would be written, no file changes
    python goodwe_automation.py --refresh    → clear cache and re-scan all stations
"""

import re
import sys
import json
import time
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy

# ─────────────────────────────────────────────
#  CONFIGURATION — edit these before running
# ─────────────────────────────────────────────
GOODWE_EMAIL    = "#########"
GOODWE_PASSWORD = "#########"
BASE_EXCEL      = r"C:\Users\Atheef Brantel Lanka\Desktop\24 March 2026\Aftersales_Maintenance_2nd_Service_Updated.xlsx"
SHEET_NAME      = "2nd Service "
OFFLINE_SHEET   = "Offline this Week"
OUTPUT_FOLDER   = r"C:\Users\Atheef Brantel Lanka\Desktop\24 March 2026"
# ─────────────────────────────────────────────

LOGIN_URL      = "https://www.semsportal.com/api/v2/Common/CrossLogin"
STATIONS_URL   = "https://hk-gateway.semsportal.com/web/sems/sems-plant/api/portal/stations/page"
DEVICES_URL    = "https://hk-gateway.semsportal.com/web/sems/sems-plant/api/stations/device/all-status"

CACHE_FILE  = Path(__file__).parent / "sn_cache.json"

# Excel column indices (0-based for pandas)
COL_BL         = 0
COL_NAME       = 1
COL_STATUS     = 2   # column C
COL_SN         = 3
COL_DAYS_START = 11  # first day column (Sunday)

MAX_RETRIES = 5
RETRY_DELAY = 5
PAGE_SIZE   = 50

# ── Inverter status codes
STATUS_ONLINE   = {5, 6, 7}   # green
STATUS_WAITING  = {3, 8}      # yellow
STATUS_OFFLINE  = {0, 2, 4}   # red

# Exact colours matched from your existing Excel file
FILL_GREEN  = PatternFill("solid", fgColor="FF66FF66")
FILL_YELLOW = PatternFill("solid", fgColor="FFFFFF00")
FILL_RED    = PatternFill("solid", fgColor="FFFF0000")
FILL_CLEAR  = PatternFill("none")

FONT_BLACK  = Font(color="FF000000")


# ══════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════

def parse_week_from_header(df):
    cell = str(df.iloc[0, COL_DAYS_START])
    m = re.search(r'\((\d+)-(\d+)/(\w+)/(\d{4})\)', cell)
    if not m:
        raise ValueError(
            f"Cannot parse date range from: '{cell}'\n"
            f"Expected format: Weekly Generation(15-21/March/2026)(kWh)"
        )
    months = {
        "January":1,"February":2,"March":3,"April":4,"May":5,"June":6,
        "July":7,"August":8,"September":9,"October":10,"November":11,"December":12
    }
    month = months.get(m.group(3))
    if not month:
        raise ValueError(f"Unknown month: '{m.group(3)}'")
    return (
        datetime(int(m.group(4)), month, int(m.group(1))),
        datetime(int(m.group(4)), month, int(m.group(2)))
    )

def parse_serial_numbers(raw):
    skip = {"nan", "", "no details", "no inverter show in an invoice"}
    if not raw or str(raw).strip().lower() in skip:
        return []
    parts = re.split(r'[,\n]+', str(raw).strip())
    return [p.strip() for p in parts if p.strip() and len(p.strip()) > 4]

def make_output_filename(start_date, end_date):
    folder = Path(OUTPUT_FOLDER)
    folder.mkdir(parents=True, exist_ok=True)
    name = f"Generation_{start_date.day}-{end_date.day}_{start_date.strftime('%B_%Y')}.xlsx"
    return folder / name

def status_to_label_and_fill(status_code):
    if status_code in STATUS_ONLINE:
        return "ONLINE",  FILL_GREEN,  FONT_BLACK
    elif status_code in STATUS_WAITING:
        return "WAITING", FILL_YELLOW, FONT_BLACK
    elif status_code in STATUS_OFFLINE:
        return "OFFLINE", FILL_RED,    FONT_BLACK
    return None, FILL_CLEAR, FONT_BLACK

def safe_write_cell(sheet, row_idx, col_idx, val):
    """Safely writes a value to a cell, automatically bypassing MergedCell read-only errors."""
    try:
        sheet.cell(row=row_idx, column=col_idx).value = val
    except AttributeError:
        # If it's a MergedCell, find the master cell of the merge and update that instead
        for m_range in sheet.merged_cells.ranges:
            if sheet.cell(row=row_idx, column=col_idx).coordinate in m_range:
                sheet.cell(row=m_range.min_row, column=m_range.min_col).value = val
                break


# ══════════════════════════════════════════════
#  CACHE
# ══════════════════════════════════════════════

def load_cache():
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, "r") as f:
                data = json.load(f)
            goodwe  = sum(1 for v in data.values() if v is not None)
            ignored = sum(1 for v in data.values() if v is None)
            print(f"  Cache loaded: {goodwe} GoodWe SNs, {ignored} non-GoodWe SNs.")
            return data
        except Exception:
            print("  Cache corrupted — rebuilding.")
    return {}

def save_cache(mapping):
    with open(CACHE_FILE, "w") as f:
        json.dump(mapping, f, indent=2)
    goodwe  = sum(1 for v in mapping.values() if v is not None)
    ignored = sum(1 for v in mapping.values() if v is None)
    print(f"  Cache saved: {goodwe} GoodWe + {ignored} non-GoodWe → {CACHE_FILE.name}")

def clear_cache():
    if CACHE_FILE.exists():
        CACHE_FILE.unlink()
        print("  Cache cleared.")
    else:
        print("  No cache file found.")


# ══════════════════════════════════════════════
#  AUTHENTICATION
# ══════════════════════════════════════════════

def login(email, password):
    print("Logging in...")
    session = requests.Session()
    session.headers.update({
        "Content-Type": "application/json",
        "Token": json.dumps({"version":"v2.1.0","client":"web","language":"en"}),
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    })

    resp = _api_call(session, "POST", LOGIN_URL, json={"account": email, "pwd": password})
    data = resp.json()

    if data.get("hasError") or data.get("msg","").lower() not in ("successful",""):
        raise ValueError(f"Login failed: {data.get('msg')}\nResponse: {data}")

    td = data["data"]
    session.headers.update({
        "Token": json.dumps({
            "version":   td.get("version", "v2.1.0"),
            "client":    td.get("client", "web"),
            "language":  td.get("language", "en"),
            "timestamp": td.get("timestamp", 0),
            "uid":       td.get("uid", ""),
            "token":     td.get("token", ""),
        })
    })

    session._credentials  = (email, password)
    session._token_expiry = time.time() + 3600
    print("  Login Successful🥰.")
    return session

def ensure_session(session):
    if time.time() > session._token_expiry - 300:
        print("  Token expiring — refreshing login...")
        email, password = session._credentials
        new_session = login(email, password)
        session.headers.update({"Token": new_session.headers["Token"]})
        session._token_expiry = new_session._token_expiry


# ══════════════════════════════════════════════
#  API CALL WITH SMART RETRY
# ══════════════════════════════════════════════

def _api_call(session, method, url, **kwargs):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.request(method, url, timeout=30, **kwargs)
            
            # Catch Rate Limiting (HTTP 429) explicitly and back off
            if resp.status_code == 429:
                wait_time = 15 * attempt  # Exponential cool-down: 15s, 30s, 45s, 60s, 75s
                print(f"    [Rate Limited] Server cooling down... waiting {wait_time}s (Attempt {attempt}/{MAX_RETRIES})")
                time.sleep(wait_time)
                if attempt == MAX_RETRIES:
                    resp.raise_for_status()
                continue
                
            resp.raise_for_status()
            return resp
        except Exception as e:
            if attempt == MAX_RETRIES:
                raise
            print(f"    Attempt {attempt} failed ({e}) — retrying in {RETRY_DELAY}s...")
            time.sleep(RETRY_DELAY)


# ══════════════════════════════════════════════
#  STATION + SN MAPPING
# ══════════════════════════════════════════════

def get_all_station_ids(session):
    print("  Fetching all station IDs (may take ~1 min)...")
    all_ids = []
    page    = 1
    while True:
        ensure_session(session)
        resp  = _api_call(session, "POST", STATIONS_URL, json={"current": page, "size": PAGE_SIZE})
        data  = resp.json()
        items = data.get("data", {}).get("dataList", [])
        total = data.get("data", {}).get("total", 0)

        for item in items:
            sid = item.get("id", "").strip()
            if sid:
                all_ids.append(sid)

        print(f"    Page {page}: {len(all_ids)}/{total} loaded", end="\r")
        if len(all_ids) >= total or not items:
            break
        page += 1
        time.sleep(2)

    print(f"\n  Total stations loaded: {len(all_ids)}")
    return all_ids

def get_device_info_for_station(session, station_id):
    try:
        resp = _api_call(session, "GET", DEVICES_URL, params={"stationId": station_id})
        data = resp.json()
        if data.get("code") != "00000":
            return {}

        result = {}
        for device in data.get("data", {}).get("deviceDetailList", []):
            if device.get("deviceType") == "INVERTER":
                for status_entry in device.get("statusDetailList", []):
                    status_code = status_entry.get("status")
                    for sn in status_entry.get("snList", []):
                        sn = sn.strip()
                        if sn:
                            result[sn] = status_code
        return result
    except Exception:
        return {}

def build_sn_to_station_map(session, station_ids, target_sns, existing_cache):
    mapping   = dict(existing_cache)
    remaining = set(sn for sn in target_sns if sn not in mapping)
    total     = len(station_ids)

    if not remaining:
        print("  All SNs already cached — skipping station scan.")
        return mapping

    print(f"  Scanning {total} stations for {len(remaining)} uncached SNs...")
    for i, sid in enumerate(station_ids):
        if not remaining:
            break

        ensure_session(session)
        device_info = get_device_info_for_station(session, sid)

        for sn, status_code in device_info.items():
            if sn in remaining:
                mapping[sn] = {"id": sid, "status": status_code}
                remaining.discard(sn)
                print(f"  FOUND: {sn} → status {status_code}  ({len(remaining)} remaining)")

        if (i + 1) % 100 == 0:
            print(f"  Checked {i+1}/{total}, found {sum(1 for v in mapping.values() if v)} GoodWe SNs...")
        time.sleep(2)

    if remaining:
        print(f"\n  {len(remaining)} SNs not found → marking as non-GoodWe:")
        for sn in sorted(remaining):
            mapping[sn] = None
    return mapping

def refresh_statuses(session, mapping):
    print("\nRefreshing live inverter statuses...")
    station_to_sns = {}
    for sn, info in mapping.items():
        if info is None: continue
        sid = info.get("id")
        if sid: station_to_sns.setdefault(sid, []).append(sn)

    updated = 0
    for sid, sns in station_to_sns.items():
        ensure_session(session)
        device_info = get_device_info_for_station(session, sid)
        for sn in sns:
            if sn in device_info:
                mapping[sn]["status"] = device_info[sn]
                updated += 1
        time.sleep(2) # Increased delay to prevent 429 errors

    print(f"  Status refreshed for {updated} inverters.")
    return mapping


# ══════════════════════════════════════════════
#  GENERATION DATA
# ══════════════════════════════════════════════

def get_station_daily_generation(session, station_id, start_date, end_date):
    ensure_session(session)
    v2_url = "https://www.semsportal.com/api/v2/PowerStationMonitor/GetPowerStationPowerAndIncomeByDay"
    results = {}
    day_count = (end_date - start_date).days + 1
    
    payload = {
        "id": station_id,
        "powerstation_id": station_id,
        "powerStationId": station_id,
        "date": end_date.strftime("%Y-%m-%d"), 
        "count": day_count
    }
    try:
        resp = _api_call(session, "POST", v2_url, json=payload)
        data = resp.json()
        code = str(data.get("code"))
        
        if code in ("00000", "0"):
            for item in data.get("data", []):
                raw_date = str(item.get("d", ""))
                try:
                    date_obj = pd.to_datetime(raw_date)
                    date_str = date_obj.strftime("%Y-%m-%d")
                except Exception:
                    continue
                val = item.get("p")
                if date_str and val is not None:
                    if start_date <= date_obj <= end_date:
                        results[date_str] = round(float(val), 1)
            if results: return results
    except Exception:
        pass 

    for i in range(day_count):
        current = start_date + timedelta(days=i)
        single_payload = {
            "id": station_id, "powerstation_id": station_id,
            "powerStationId": station_id, "date": current.strftime("%Y-%m-%d"), "count": 1
        }
        try:
            resp = _api_call(session, "POST", v2_url, json=single_payload)
            data = resp.json()
            if str(data.get("code")) in ("00000", "0"):
                for item in data.get("data", []):
                    raw_date = str(item.get("d", current.strftime("%Y-%m-%d")))
                    try: date_str = pd.to_datetime(raw_date).strftime("%Y-%m-%d")
                    except Exception: date_str = current.strftime("%Y-%m-%d")
                    val = item.get("p")
                    if val is not None: results[date_str] = round(float(val), 1)
            time.sleep(1.5) # Increased delay to prevent 429 errors
        except Exception:
            pass
            
    if not results:
        print(f"      → Unreachable")
    return results

def get_week_generation(session, sns, sn_to_station, start_date, end_date):
    day_count     = (end_date - start_date).days + 1
    all_days      = {}
    seen_stations = set()
    any_found     = False

    for sn in sns:
        info = sn_to_station.get(sn)
        if not info: continue
        sid = info.get("id") if isinstance(info, dict) else None
        if not sid or sid in seen_stations: continue

        seen_stations.add(sid)
        daily = get_station_daily_generation(session, sid, start_date, end_date)
        if daily:
            any_found = True
            for date_str, kwh in daily.items():
                all_days[date_str] = round(all_days.get(date_str, 0.0) + kwh, 1)
        time.sleep(1.5) # Increased delay to prevent 429 errors

    result = []
    for i in range(day_count):
        d   = (start_date + timedelta(days=i)).strftime("%Y-%m-%d")
        val = all_days.get(d)
        if val is not None: result.append(val)
        elif any_found: result.append(0.0)
        else: result.append(None)
    return result

def get_station_status_for_sns(sns, sn_to_station):
    codes = set()
    for sn in sns:
        info = sn_to_station.get(sn)
        if isinstance(info, dict) and "status" in info:
            codes.add(info["status"])

    if not codes: return None
    if codes & STATUS_ONLINE: return max(codes & STATUS_ONLINE)
    if codes & STATUS_WAITING: return 3
    if codes & STATUS_OFFLINE: return 0
    return None


# ══════════════════════════════════════════════
#  OUTPUT EXCEL
# ══════════════════════════════════════════════

def build_day_headers(start_date, end_date):
    day_names = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    headers   = []
    current   = start_date
    while current <= end_date:
        headers.append(f"{current.day}/{day_names[current.weekday()]}")
        current += timedelta(days=1)
    return headers

def create_output_excel(base_path, output_path, sheet_name,
                        start_date, end_date,
                        generation_data, status_data, offline_rows):
    wb = load_workbook(base_path)
    ws = wb[sheet_name]

    # 1. Update generation header safely
    month_name = start_date.strftime("%B")
    generation_header_str = f"Weekly Generation({start_date.day}-{end_date.day}/{month_name}/{start_date.year})(kWh)"
    safe_write_cell(ws, 1, COL_DAYS_START + 1, generation_header_str)

    # 2. Update day sub-headers safely
    for i, hdr in enumerate(build_day_headers(start_date, end_date)):
        safe_write_cell(ws, 2, COL_DAYS_START + 1 + i, hdr)

    # 3. Clear all existing generation data
    for row in range(3, ws.max_row + 1):
        for col_offset in range(7):
            safe_write_cell(ws, row, COL_DAYS_START + 1 + col_offset, None)

    # 4. Write generation values
    for excel_row, values in generation_data:
        for day_offset, val in enumerate(values):
            if val is not None:
                safe_write_cell(ws, excel_row, COL_DAYS_START + 1 + day_offset, val)

    # 5. Write status + colour
    STATUS_COL = COL_STATUS + 1
    for excel_row, status_code in status_data:
        cell  = ws.cell(row=excel_row, column=STATUS_COL)
        label, fill, font = status_to_label_and_fill(status_code)
        if label is not None:
            try:
                cell.value = label
                cell.fill  = fill
                cell.font  = font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            except AttributeError:
                pass # Skip styling if strictly merged

    # ────────────────────────────────────────────────────────
    # 6. COPY OFFLINE ROWS TO DEDICATED SHEET
    # ────────────────────────────────────────────────────────
    if OFFLINE_SHEET in wb.sheetnames:
        off_ws = wb[OFFLINE_SHEET]

        # Update headers in the Offline sheet safely
        safe_write_cell(off_ws, 1, COL_DAYS_START + 1, generation_header_str)
        for i, hdr in enumerate(build_day_headers(start_date, end_date)):
            safe_write_cell(off_ws, 2, COL_DAYS_START + 1 + i, hdr)

        # Clear existing data in Offline sheet (row 3 downwards)
        if off_ws.max_row >= 3:
            off_ws.delete_rows(3, off_ws.max_row - 2)

        # Copy the rows marked as offline
        target_row = 3
        for src_row_idx in offline_rows:
            off_ws.row_dimensions[target_row].height = ws.row_dimensions[src_row_idx].height
            
            for col_idx in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=src_row_idx, column=col_idx)
                tgt_cell = off_ws.cell(row=target_row, column=col_idx)
                
                try:
                    tgt_cell.value = src_cell.value
                    if src_cell.has_style:
                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.number_format = copy(src_cell.number_format)
                        tgt_cell.protection = copy(src_cell.protection)
                        tgt_cell.alignment = copy(src_cell.alignment)
                except AttributeError:
                    continue # Skip read-only merged child cells
                    
            target_row += 1

    wb.save(output_path)


# ══════════════════════════════════════════════
#  SUMMARY
# ══════════════════════════════════════════════

def print_summary(results, output_path, preview, offline_rows_count):
    updated   = [r for r in results if r["status"] == "updated"]
    offline   = [r for r in results if r["status"] == "offline"]
    waiting   = [r for r in results if r["status"] == "waiting"]
    not_found = [r for r in results if r["status"] == "not_found"]
    non_gw    = [r for r in results if r["status"] == "non_goodwe"]
    skipped   = [r for r in results if r["status"] == "skipped"]

    print("\n" + "═" * 55)
    print("  SUMMARY")
    print("═" * 55)
    print(f"  Data written (ONLINE)      : {len(updated)}")
    print(f"  Waiting / standby          : {len(waiting)}")
    print(f"  Offline / zero generation  : {len(offline)}")
    print(f"  Not found in portal        : {len(not_found)}")
    print(f"  Non-GoodWe inverters       : {len(non_gw)}")
    print(f"  Skipped (no S/N)           : {len(skipped)}")
    print(f"  --> Copied {offline_rows_count} offline rows to '{OFFLINE_SHEET}' sheet.")
    print("═" * 55)

    if not preview and output_path:
        print(f"\n  Output: {Path(output_path).name}")

    if not_found:
        print("\n  Not found in portal:")
        for r in not_found:
            print(f"    [{r['bl']}] {r['name']}")
    if offline:
        print("\n  Offline / zero this week:")
        for r in offline:
            print(f"    [{r['bl']}] {r['name']}")


# ══════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════

def main():
    preview = "--preview" in sys.argv
    refresh = "--refresh" in sys.argv

    print("═" * 55)
    print("  GoodWe Weekly Generation Automation  v12")
    if preview:
        print("  *** PREVIEW MODE — no files will be modified ***")
    if refresh:
        print("  *** REFRESH MODE — cache will be rebuilt ***")
    print("═" * 55)

    if refresh:
        print("\nClearing cache...")
        clear_cache()

    print(f"\nReading base file: {BASE_EXCEL}")
    df = pd.read_excel(BASE_EXCEL, sheet_name=SHEET_NAME, header=None)
    start_date, end_date = parse_week_from_header(df)
    print(f"Week: {start_date.strftime('%d %B %Y')} → {end_date.strftime('%d %B %Y')}")

    all_target_sns = set()
    for idx in range(2, len(df)):
        all_target_sns.update(parse_serial_numbers(df.iloc[idx][COL_SN]))
    print(f"Total SNs in Excel: {len(all_target_sns)}")

    session = login(GOODWE_EMAIL, GOODWE_PASSWORD)

    print("\nLoading SN cache...")
    cached_map = load_cache()

    uncached = [sn for sn in all_target_sns if sn not in cached_map]
    if uncached:
        print(f"  {len(uncached)} uncached SNs — scanning stations...")
        station_ids   = get_all_station_ids(session)
        sn_to_station = build_sn_to_station_map(session, station_ids, all_target_sns, cached_map)
        if not preview: save_cache(sn_to_station)
    else:
        print("  All SNs cached — skipping station scan. ✓")
        sn_to_station = cached_map

    sn_to_station = refresh_statuses(session, sn_to_station)
    if not preview: save_cache(sn_to_station)

    goodwe_count = sum(1 for v in sn_to_station.values() if v is not None)
    print(f"  GoodWe inverters to process: {goodwe_count}")

    generation_data = []   
    status_data     = []   
    results         = []
    offline_rows    = []   

    print(f"\nFetching weekly generation...\n")

    for idx in range(2, len(df)):
        row      = df.iloc[idx]
        customer = str(row[COL_NAME]).strip() if pd.notna(row[COL_NAME]) else ""
        bl       = str(row[COL_BL]).strip()   if pd.notna(row[COL_BL])   else ""
        raw_sn   = row[COL_SN]

        if not customer or customer == "nan": continue

        sns = parse_serial_numbers(raw_sn)
        if not sns:
            print(f"  [{bl}] {customer} — skipped (no S/N)")
            results.append({"bl": bl, "name": customer, "status": "skipped"})
            continue

        goodwe_sns = [sn for sn in sns if isinstance(sn_to_station.get(sn), dict)]
        non_gw_sns = [sn for sn in sns if sn in sn_to_station and sn_to_station[sn] is None]

        if not goodwe_sns and non_gw_sns:
            print(f"  [{bl}] {customer} — non-GoodWe, skipped")
            results.append({"bl": bl, "name": customer, "status": "non_goodwe"})
            continue
        if not goodwe_sns:
            print(f"  [{bl}] {customer} — not found in portal")
            results.append({"bl": bl, "name": customer, "status": "not_found"})
            continue

        status_code = get_station_status_for_sns(sns, sn_to_station)
        
        print(f"  [{bl}] {customer}")
        daily_values = get_week_generation(session, sns, sn_to_station, start_date, end_date)

        excel_row = idx + 1   

        if status_code is not None:
            status_data.append((excel_row, status_code))

        is_offline = False

        if all(v is None for v in daily_values):
            print(f"    → Unreachable")
            results.append({"bl": bl, "name": customer, "status": "not_found"})
        elif all(v == 0.0 or v is None for v in daily_values):
            label = "waiting" if status_code in STATUS_WAITING else "offline"
            print(f"    → {label.capitalize()} / zero: {daily_values}")
            results.append({"bl": bl, "name": customer, "status": label})
            generation_data.append((excel_row, daily_values))
            if label == "offline": 
                is_offline = True
        else:
            print(f"    → {daily_values}")
            results.append({"bl": bl, "name": customer, "status": "updated"})
            generation_data.append((excel_row, daily_values))
            if status_code in STATUS_OFFLINE:
                is_offline = True

        if is_offline:
            offline_rows.append(excel_row)

    output_path = None
    if not preview:
        output_path = make_output_filename(start_date, end_date)
        print(f"\nCreating output file: {Path(output_path).name} ...")
        create_output_excel(
            BASE_EXCEL, output_path, SHEET_NAME,
            start_date, end_date,
            generation_data, status_data, offline_rows
        )
        print(f"  Output saved. ✓")
        print(f"  Base file untouched. ✓")
    else:
        print(f"\n[PREVIEW] Would create: Generation_{start_date.day}-{end_date.day}_{start_date.strftime('%B_%Y')}.xlsx")
        print(f"[PREVIEW] Would write {len(generation_data)} generation rows.")
        print(f"[PREVIEW] Would update {len(status_data)} status cells.")
        print(f"[PREVIEW] Would copy {len(offline_rows)} offline systems to '{OFFLINE_SHEET}' sheet.")
        print("[PREVIEW] No files modified.")

    print_summary(results, output_path, preview, len(offline_rows))

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
        sys.exit(0)
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)